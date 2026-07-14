"""Extrai seed-data.json da PLANILHA DE ESTOQUE.xlsx (valores + lógica Fábrica)."""
import json
import re
from datetime import date
from pathlib import Path

import openpyxl

XLSX = Path(r"c:\Users\Mauro\Downloads\PLANILHA DE ESTOQUE.xlsx")
OUT = Path(r"c:\Users\Mauro\Documents\Obra 103\seed-data.json")
SEED_VERSION = date.today().isoformat() + "-filtros"  # força refresh no app quando muda

wb = openpyxl.load_workbook(XLSX, data_only=True)
wb_f = openpyxl.load_workbook(XLSX, data_only=False)


def slug(s):
    s = (s or "").strip().lower()
    repl = {
        "á": "a", "à": "a", "ã": "a", "â": "a", "é": "e", "ê": "e",
        "í": "i", "ó": "o", "ô": "o", "õ": "o", "ú": "u", "ç": "c", "ü": "u",
    }
    for a, b in repl.items():
        s = s.replace(a, b)
    return re.sub(r"[^a-z0-9]+", "-", s).strip("-") or "x"


def num(v):
    try:
        if v is None or str(v).strip() == "":
            return 0.0
        return float(v)
    except Exception:
        return 0.0


def as_date(v):
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        try:
            return v.strftime("%Y-%m-%d")
        except Exception:
            return ""
    s = str(v).strip()
    if s in ("True", "False", "#VALUE!", "#REF!", ""):
        return ""
    return s[:10]


def as_bool(v):
    if v is True or v is False:
        return bool(v)
    s = str(v or "").strip().upper()
    return s in ("OK", "X", "SIM", "S", "TRUE", "1", "✓", "✔")


def empty_estoque():
    return {
        "saldo": 0.0,
        "minimo": 0.0,
        "envio": 0.0,
        "validade1": "",
        "validade2": "",
        "validade3": "",
        "okEntregador": False,
        "okLoja": False,
    }


# ── Produtos (catálogo = ESTOQUE CENTRAL) ──
ws = wb["ESTOQUE CENTRAL"]
produtos = []
seen = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    cc, cat, nome, un = row[0], row[1], row[2], row[3]
    if not nome or str(nome).strip().upper() in ("INATIVO", ""):
        continue
    nome = str(nome).strip()
    cat = str(cat or "").strip()
    cc = str(cc or "").strip()
    un = str(un or "UN").strip()
    ativo = cc.upper() != "INATIVO" and cat.upper() != "INATIVO"
    key = nome.upper()
    if key in seen:
        continue
    seen.add(key)
    pid = slug(nome)[:48]
    base = pid
    n = 2
    while any(p["id"] == pid for p in produtos):
        pid = f"{base}-{n}"
        n += 1
    produtos.append({
        "id": pid,
        "centroCusto": cc,
        "categoria": cat,
        "nome": nome,
        "unidade": un,
        "ativo": ativo,
    })

by_name = {p["nome"].upper(): p["id"] for p in produtos}

# ── Estoque lojas ──
# CENTRAL: saldo, min, val1, val2, val3, envio emergência
# LOJAS:   saldo, min, envio, val1, val2, val3, okEntregador, okLoja
loja_sheets = {
    "ESTOQUE CENTRAL": "central",
    "LAGO SUL": "lago-sul",
    "AGUAS CLARAS": "aguas-claras",
    "ASA NORTE": "asa-norte",
    "ASA SUL": "asa-sul",
}

estoques = {lid: {} for lid in loja_sheets.values()}
for sheet, lid in loja_sheets.items():
    ws = wb[sheet]
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome = row[2]
        if not nome:
            continue
        pid = by_name.get(str(nome).strip().upper())
        if not pid:
            continue
        entry = empty_estoque()
        entry["saldo"] = num(row[4])
        entry["minimo"] = num(row[5])
        if lid == "central":
            entry["validade1"] = as_date(row[6] if len(row) > 6 else None)
            entry["validade2"] = as_date(row[7] if len(row) > 7 else None)
            entry["validade3"] = as_date(row[8] if len(row) > 8 else None)
            entry["envio"] = num(row[9] if len(row) > 9 else 0)
        else:
            entry["envio"] = num(row[6] if len(row) > 6 else 0)
            entry["validade1"] = as_date(row[7] if len(row) > 7 else None)
            entry["validade2"] = as_date(row[8] if len(row) > 8 else None)
            entry["validade3"] = as_date(row[9] if len(row) > 9 else None)
            entry["okEntregador"] = as_bool(row[10] if len(row) > 10 else None)
            entry["okLoja"] = as_bool(row[11] if len(row) > 11 else None)
        estoques[lid][pid] = entry

# ── Cotações ──
# PERSONALIZADOS na planilha NÃO é aba de disputa (Amorix/Oesa/…): é "COMPRA ITENS
# PERSONALIZADOS" com layout próprio. Itens da categoria PERSONALIZADOS ficam ocultos
# nas outras abas de cotação/RESULTADO — só o fornecedor `personalizados` os cota.
forn_sheets = {
    "AMORIX": "amorix",
    "OESA": "oesa",
    "GARRA": "garra",
    "REI NOSSO": "rei-nosso",
}
cotacoes = {fid: {} for fid in list(forn_sheets.values()) + ["personalizados"]}
for sheet, fid in forn_sheets.items():
    if sheet not in wb.sheetnames:
        continue
    ws = wb[sheet]
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome = row[1]
        if not nome:
            continue
        pid = by_name.get(str(nome).strip().upper())
        if not pid:
            continue
        # Não importa preço de itens PERSONALIZADOS nas abas dos outros fornecedores
        prod = next((p for p in produtos if p["id"] == pid), None)
        if prod and "PERSONAL" in str(prod.get("categoria") or "").upper():
            continue
        status = str(row[6] or "FALTA").strip().upper() if len(row) > 6 and row[6] else "FALTA"
        if status not in ("OK", "FALTA"):
            status = "FALTA" if not num(row[4]) else "OK"
        cotacoes[fid][pid] = {
            "qtde": num(row[3]),
            "preco": num(row[4]),
            "observacoes": str(row[5] or ""),
            "status": status,
        }

# Aba PERSONALIZADOS: ITEM | MEDIDAS | ESPEC | GRAM | TIRAGEM | VR UNIT | …
# Nomes curtos da aba → id do catálogo (categoria PERSONALIZADOS).
PERSONALIZADOS_MAP = {
    "saco kraft g": "saco-kraft-g-24x32x15-5",
    "saco kraft p": "saco-kraft-p-18x31x9",
    "box": "emb-box-chines-500ml-branca-kraft",
    "adesivo box": "adesivo-box-conferencia-e-lacre",
    "caixa batata": "caixa-batata-frita-embalagem-m",
    "lacre seg": "lacre-p-saco-kraft",
    "lacre marmita": "sacola-38-x-48-ou-lacre-marmita",
}

def _eval_preco(v):
    if v is None or str(v).strip() == "":
        return 0.0
    if isinstance(v, str) and v.strip().startswith("="):
        expr = v.strip().lstrip("=").replace(" ", "")
        if re.fullmatch(r"\d+(?:\.\d+)?/\d+(?:\.\d+)?", expr):
            a, b = expr.split("/")
            return float(a) / float(b)
        return 0.0
    return num(v)

if "PERSONALIZADOS" in wb.sheetnames:
    ws = wb["PERSONALIZADOS"]
    for row in ws.iter_rows(min_row=5, values_only=True):
        item = row[0] if row else None
        if not item:
            continue
        pid = PERSONALIZADOS_MAP.get(str(item).strip().lower())
        if not pid or not any(p["id"] == pid for p in produtos):
            continue
        preco = _eval_preco(row[5] if len(row) > 5 else 0)
        if preco <= 0:
            continue
        cotacoes["personalizados"][pid] = {
            "qtde": 0.0,
            "preco": round(preco, 6),
            "observacoes": " | ".join(
                str(x).strip() for x in (row[1], row[2], row[3], row[4]) if x and str(x).strip()
            ),
            "status": "OK",
        }

# ── Produção (aba PRODUÇÃO) ──
producao = {}
prod_sheet = [n for n in wb.sheetnames if "PRODU" in n.upper()][0]
ws = wb[prod_sheet]
for row in ws.iter_rows(min_row=2, values_only=True):
    nome = row[2]
    if not nome:
        continue
    pid = by_name.get(str(nome).strip().upper())
    if not pid:
        continue
    lista = str(row[6] or "").strip().upper()
    producao[pid] = {
        "lista": "PRODUZIR" if "PRODUZIR" in lista else "",
        "totalProduzir": num(row[7]),
        "qtdeBaldes": num(row[8]),
        "totalProduzido": num(row[9]),
    }

# ── Fábrica: estoque + receitas / auto produção (fórmulas) ──
ws_fab_v = wb["FABRICA"]
ws_fab = wb_f["FABRICA"]
wst = wb_f["TOTAIS"]

row_id = {}
for r in range(2, wst.max_row + 1):
    nome = wst.cell(r, 3).value
    if not nome:
        continue
    pid = by_name.get(str(nome).strip().upper())
    if not pid:
        pid = slug(str(nome))
    row_id[r] = pid


def subst_h(formula):
    if not formula:
        return None
    f = formula.strip()
    if f.startswith("="):
        f = f[1:]

    def repl(m):
        rr = int(m.group(1))
        return f"Q['{row_id.get(rr) or '0'}']" if row_id.get(rr) else "0"

    return re.sub(r"H(\d+)", repl, f)


receitas = {}
producao_flags = {}
baldes = {}
estoque_fab = {}

for r, pid in row_id.items():
    e = ws_fab.cell(r, 5).value
    f = ws_fab.cell(r, 6).value
    g = ws_fab.cell(r, 7).value
    h = ws_fab.cell(r, 8).value
    i = ws_fab.cell(r, 9).value
    j = ws_fab.cell(r, 10).value
    k = ws_fab.cell(r, 11).value
    ev = ws_fab_v.cell(r, 5).value
    fv = ws_fab_v.cell(r, 6).value

    fab_entry = empty_estoque()
    fab_entry["saldo"] = float(ev or 0) if not isinstance(ev, str) else 0.0
    fab_entry["minimo"] = float(fv or 0) if not isinstance(fv, str) else 0.0
    fab_entry["minimoAuto"] = False
    estoque_fab[pid] = fab_entry

    if isinstance(f, str) and f.startswith("="):
        fs = f[1:].replace(" ", "")
        if re.fullmatch(r"I\d+\*J\d+\+K\d+", fs) or re.fullmatch(r"I\d+\*J\d+", fs):
            i_expr = subst_h(i if isinstance(i, str) else str(i or 0))
            fator = float(j or 0) if not isinstance(j, str) else 0.0
            margem = float(k or 0) if not isinstance(k, str) else 0.0
            expr = f"({i_expr})*{fator}+{margem}"
        else:
            expr = subst_h(f)
        receitas[pid] = {"expr": expr, "autoMinimo": True}
        fab_entry["minimoAuto"] = True

    hard_produzir = isinstance(g, str) and g.strip().upper() == "PRODUZIR"
    has_totais_g = isinstance(h, str) and "TOTAIS!G" in h.upper().replace(" ", "")
    if_formula = isinstance(g, str) and "PRODUZIR" in g.upper()
    if hard_produzir or has_totais_g or if_formula:
        producao_flags[pid] = {
            "auto": True,
            "sempreLista": hard_produzir and not if_formula,
        }
        if pid not in producao:
            producao[pid] = {
                "lista": "",
                "totalProduzir": 0.0,
                "qtdeBaldes": 0.0,
                "totalProduzido": 0.0,
            }
        producao[pid]["auto"] = True

    if isinstance(i, str) and i.startswith("=") and re.search(r"H\d+/", i.replace(" ", "")):
        m = re.search(r"H(\d+)/(\d+(?:\.\d+)?)", i.replace(" ", ""))
        if m and int(m.group(1)) == r:
            baldes[pid] = float(m.group(2))

# Garante entrada de fábrica para todos os produtos do catálogo
for p in produtos:
    if p["id"] not in estoque_fab:
        estoque_fab[p["id"]] = empty_estoque()
        estoque_fab[p["id"]]["minimoAuto"] = False

estoques["fabrica"] = estoque_fab

# ── Produtos visíveis por loja (filtro Excel = linhas NÃO ocultas) ──
produtos_por_loja = {}
vis_sheets = {
    "ESTOQUE CENTRAL": "central",
    "LAGO SUL": "lago-sul",
    "AGUAS CLARAS": "aguas-claras",
    "ASA NORTE": "asa-norte",
    "ASA SUL": "asa-sul",
    "FABRICA": "fabrica",
}
for sheet, lid in vis_sheets.items():
    ws = wb[sheet]
    ids = []
    seen_vis = set()
    for i in range(2, (ws.max_row or 1) + 1):
        if ws.row_dimensions[i].hidden:
            continue
        nome = ws.cell(i, 3).value
        if not nome:
            continue
        pid = by_name.get(str(nome).strip().upper())
        if not pid or pid in seen_vis:
            continue
        seen_vis.add(pid)
        ids.append(pid)
    produtos_por_loja[lid] = ids

out = {
    "seedVersion": SEED_VERSION,
    "produtos": produtos,
    "estoques": estoques,
    "cotacoes": cotacoes,
    "producao": producao,
    "receitasMinimoFabrica": receitas,
    "producaoAuto": producao_flags,
    "baldesPor": baldes,
    "produtosPorLoja": produtos_por_loja,
}

OUT.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
OUT_JS = OUT.with_suffix(".js")
OUT_JS.write_text("window.SEED_DATA = " + json.dumps(out, ensure_ascii=False, separators=(",", ":")) + ";\n", encoding="utf-8")

print("seedVersion", SEED_VERSION)
print("produtos", len(produtos), "ativos", sum(1 for p in produtos if p["ativo"]))
for lid, m in estoques.items():
    nz = sum(1 for e in m.values() if e.get("saldo") or e.get("minimo") or e.get("envio"))
    print(f"  estoque {lid}: {len(m)} itens ({nz} com valores)")
for lid, ids in produtos_por_loja.items():
    print(f"  visiveis {lid}: {len(ids)}")
for fid, m in cotacoes.items():
    ok = sum(1 for e in m.values() if e.get("status") == "OK")
    print(f"  cotacao {fid}: {len(m)} ({ok} OK)")
print("producao", len(producao), "receitas", len(receitas), "producaoAuto", len(producao_flags), "baldes", len(baldes))
print("wrote", OUT, OUT.stat().st_size, "bytes")
print("wrote", OUT_JS, OUT_JS.stat().st_size, "bytes")
